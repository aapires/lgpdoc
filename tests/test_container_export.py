"""Tests for the safe / sensitive XLSX exports.

Critical: the safe export must NEVER include ``original_text``. The
sensitive export must include it AND only be served by its dedicated
endpoint. A regression in either direction is treated as a privacy
incident — these tests guard the boundary.
"""
from __future__ import annotations

import io
import time
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from anonymizer_api.config import Settings
from anonymizer_api.main import create_app

POLICY_PATH = Path(__file__).parent.parent / "policies" / "default.yaml"

# Synthetic original values that the seeding upload causes to be stored
# as ``original_text`` in the mapping. The privacy assertion below
# searches the XLSX bytes for these strings.
SYNTH_NAME = "Joao Silva"
SYNTH_EMAIL = "alice@example.com"


@pytest.fixture()
def api_settings(tmp_path: Path) -> Settings:
    return Settings(
        quarantine_dir=tmp_path / "quarantine",
        output_dir=tmp_path / "output",
        db_url=f"sqlite:///{tmp_path}/api.db",
        max_bytes=1 * 1024 * 1024,
        policy_path=POLICY_PATH,
        runtime_config_path=tmp_path / "runtime.json",
        use_mock_client=True,
    )


@pytest.fixture()
def api_client(api_settings: Settings) -> TestClient:
    app = create_app(api_settings)
    with TestClient(app) as client:
        yield client


def _create_container(client: TestClient, name: str = "exp") -> str:
    return client.post("/api/containers", json={"name": name}).json()[
        "container_id"
    ]


def _seed_with_raw_doc(
    client: TestClient,
    cid: str,
    *,
    body: str | None = None,
    filename: str = "doc.txt",
) -> None:
    """Sprint 5 flow: upload → pending_review → approve → ready."""
    body = body or (
        f"Cliente: {SYNTH_NAME}.\n"
        f"Email: {SYNTH_EMAIL}.\n"
    )
    files = {"file": (filename, body.encode("utf-8"), "text/plain")}
    r = client.post(f"/api/containers/{cid}/documents/raw", files=files)
    assert r.status_code == 201, r.text
    doc = r.json()

    deadline = time.monotonic() + 5.0
    while time.monotonic() < deadline:
        d = client.get(
            f"/api/containers/{cid}/documents/{doc['document_id']}"
        ).json()
        if d["status"] == "pending_review":
            break
        time.sleep(0.05)
    r = client.post(f"/jobs/{doc['job_id']}/approve", json={})
    assert r.status_code == 200
    deadline = time.monotonic() + 5.0
    while time.monotonic() < deadline:
        d = client.get(
            f"/api/containers/{cid}/documents/{doc['document_id']}"
        ).json()
        if d["status"] == "ready":
            return
        time.sleep(0.05)
    raise AssertionError("seed: doc never reached ready")


def _read_xlsx_rows(payload: bytes) -> list[list[str]]:
    wb = load_workbook(io.BytesIO(payload))
    ws = wb.active
    assert ws is not None
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c) if c is not None else "" for c in row])
    return rows


def _is_xlsx(payload: bytes) -> bool:
    # XLSX is a ZIP — first 4 bytes are PK\x03\x04.
    return payload.startswith(b"PK\x03\x04")


# ---------------------------------------------------------------------------
# Sensitive export — includes the real values, mirrors the visible mapping
# table layout (Marcador / Tipo / Valor normalizado / Valor real /
# Ocorrências / Revisão).
#
# The ``safe`` variant was deliberately removed — it never carried enough
# information to be useful in practice (the normalised value of an email
# is the email itself), so the only export now is the explicit sensitive
# one, gated behind a confirmation dialog in the UI.
# ---------------------------------------------------------------------------

class TestSensitiveExport:
    def test_returns_xlsx_with_correct_mime(
        self, api_client: TestClient
    ) -> None:
        cid = _create_container(api_client)
        _seed_with_raw_doc(api_client, cid)
        r = api_client.get(
            f"/api/containers/{cid}/mapping/export-sensitive.xlsx"
        )
        assert r.status_code == 200
        assert _is_xlsx(r.content)
        # The filename must clearly mark it as sensitive.
        assert "sensitive" in r.headers["content-disposition"].lower()

    def test_columns_match_visible_mapping_table(
        self, api_client: TestClient
    ) -> None:
        """The export's column shape must match what the operator sees
        on screen — same labels, same order, plus the ``Valor real``
        column that the visual table omits (the whole point of the
        sensitive variant)."""
        cid = _create_container(api_client)
        _seed_with_raw_doc(api_client, cid)
        r = api_client.get(
            f"/api/containers/{cid}/mapping/export-sensitive.xlsx"
        )
        rows = _read_xlsx_rows(r.content)
        headers = rows[0]
        assert headers == [
            "Marcador",
            "Tipo",
            "Valor normalizado",
            "Valor real",
            "Ocorrências",
            "Revisão",
        ]

    def test_includes_original_text_in_valor_real_column(
        self, api_client: TestClient
    ) -> None:
        cid = _create_container(api_client)
        _seed_with_raw_doc(api_client, cid)
        r = api_client.get(
            f"/api/containers/{cid}/mapping/export-sensitive.xlsx"
        )
        rows = _read_xlsx_rows(r.content)
        headers = rows[0]
        col = headers.index("Valor real")
        originals = [r[col] for r in rows[1:]]
        assert any(SYNTH_NAME in o for o in originals)

    def test_occurrences_column_lists_filenames(
        self, api_client: TestClient
    ) -> None:
        cid = _create_container(api_client)
        _seed_with_raw_doc(api_client, cid)  # uses filename "doc.txt"
        r = api_client.get(
            f"/api/containers/{cid}/mapping/export-sensitive.xlsx"
        )
        rows = _read_xlsx_rows(r.content)
        headers = rows[0]
        col = headers.index("Ocorrências")
        # Every row must list at least the seed filename.
        for body_row in rows[1:]:
            assert "doc.txt" in body_row[col]

    def test_404_for_unknown_container(self, api_client: TestClient) -> None:
        r = api_client.get(
            "/api/containers/missing/mapping/export-sensitive.xlsx"
        )
        assert r.status_code == 404


# ---------------------------------------------------------------------------
# Safe-export endpoint must NOT exist anymore
# ---------------------------------------------------------------------------

class TestSafeExportRemoved:
    def test_safe_endpoint_is_gone(self, api_client: TestClient) -> None:
        cid = _create_container(api_client)
        _seed_with_raw_doc(api_client, cid)
        r = api_client.get(
            f"/api/containers/{cid}/mapping/export-safe.xlsx"
        )
        assert r.status_code == 404


# ---------------------------------------------------------------------------
# Container isolation — exports never spill across containers
# ---------------------------------------------------------------------------

class TestExportIsolation:
    def test_export_only_includes_own_container(
        self, api_client: TestClient
    ) -> None:
        cid_a = _create_container(api_client, name="A")
        cid_b = _create_container(api_client, name="B")
        # Seed both containers via the full upload→approve flow so the
        # mapping table actually has entries.
        _seed_with_raw_doc(
            api_client,
            cid_b,
            body="Cliente: Maria Costa.\n",
            filename="b.txt",
        )
        _seed_with_raw_doc(api_client, cid_a)

        # Sensitive export of A must NOT contain "Maria Costa".
        r = api_client.get(
            f"/api/containers/{cid_a}/mapping/export-sensitive.xlsx"
        )
        rows = _read_xlsx_rows(r.content)
        headers = rows[0]
        col = headers.index("Valor real")
        originals_a = [r[col] for r in rows[1:]]
        assert any("Joao" in o for o in originals_a)
        assert all("Maria Costa" not in o for o in originals_a)
