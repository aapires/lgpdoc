from __future__ import annotations

from pathlib import Path

from .base import BaseExtractor, UnsupportedFormatError
from ._markdown import to_markdown_table
from ..document_models import ExtractionResult


class XlsxExtractor(BaseExtractor):
    """Extracts text from XLSX files.

    Each sheet becomes a single DocumentBlock containing the sheet
    rendered as a GitHub-flavoured Markdown table. The first row is
    treated as the header row — that's the convention for Excel
    sheets used as data tables.

    Empty sheets are skipped. Trailing empty rows are dropped to
    avoid trailing ``| | | | |`` lines in the output.
    """

    supported_extensions = frozenset({".xlsx"})

    def extract(self, path: Path) -> ExtractionResult:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl is required for XLSX extraction. "
                "Install it with: pip install 'openpyxl>=3.1'"
            ) from exc

        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
        except Exception as exc:
            raise UnsupportedFormatError(
                f"{path.name!r} could not be parsed as an XLSX file: {exc}"
            ) from exc

        raw: list[tuple[int | None, str]] = []
        for sheet_index, ws in enumerate(wb.worksheets, start=1):
            rows: list[list[object]] = []
            for row in ws.iter_rows(values_only=True):
                # Stop accumulating empty rows once we've started; but
                # only KEEP rows that have at least one non-empty cell.
                if any(c is not None and str(c).strip() for c in row):
                    rows.append(list(row))
            if not rows:
                continue
            rendered = to_markdown_table(rows)
            if rendered:
                raw.append((sheet_index, rendered))

        wb.close()
        return self._build_result(raw)
